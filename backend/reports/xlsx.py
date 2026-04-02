from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def build_xlsx(*, sheet_name: str, headers: list[str], rows: Iterable[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_font = Font(bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append(list(row))

    ws.freeze_panes = "A2"

    # Auto width (simple heuristic) + wrap
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

