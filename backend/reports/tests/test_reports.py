from decimal import Decimal
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework.test import APIClient

from clients.models import Client
from inventory.models import InventoryCategory, Product
from inventory.services import stock_in, stock_out
from orders.models import Order, OrderStatus


User = get_user_model()


def _auth(client: APIClient, *, email: str, password: str) -> None:
    resp = client.post("/api/auth/token/", {"email": email, "password": password}, format="json")
    token = resp.data["access"]
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {token}")


@pytest.mark.django_db
def test_tc28_dashboard_visible_for_admin():
    admin = User.objects.create_user(email="admin@example.com", name="Admin", role=User.Role.ADMIN, password="Passw0rd123")
    api = APIClient()
    _auth(api, email=admin.email, password="Passw0rd123")
    resp = api.get("/api/reports/dashboard/")
    assert resp.status_code == 200
    assert "orders" in resp.data
    assert "by_status" in resp.data
    assert resp.data["role"] == "admin"
    assert "charts" in resp.data
    assert "urgent_orders" in resp.data


@pytest.mark.django_db
def test_dashboard_visible_for_manager_and_master():
    manager = User.objects.create_user(
        email="mgr_dash@example.com", name="Mgr", role=User.Role.MANAGER, password="Passw0rd123"
    )
    master = User.objects.create_user(
        email="m_dash@example.com", name="Master", role=User.Role.MASTER, password="Passw0rd123"
    )
    api_mgr = APIClient()
    _auth(api_mgr, email=manager.email, password="Passw0rd123")
    r1 = api_mgr.get("/api/reports/dashboard/")
    assert r1.status_code == 200
    assert r1.data["role"] == "manager"
    assert "stats" in r1.data
    assert r1.data["negotiation_orders"] is not None

    api_m = APIClient()
    _auth(api_m, email=master.email, password="Passw0rd123")
    r2 = api_m.get("/api/reports/dashboard/")
    assert r2.status_code == 200
    assert r2.data["role"] == "master"
    assert r2.data["master_orders"] is not None
    assert r2.data["low_stock"] is None


@pytest.mark.django_db
def test_tc30_orders_report_and_tc31_export_xlsx():
    admin = User.objects.create_user(email="admin@example.com", name="Admin", role=User.Role.ADMIN, password="Passw0rd123")
    client = Client.objects.create(type="person", name="Иванов", phone="+70000000000")
    accepted = OrderStatus.objects.get(code="accepted")
    from orders.models import Order

    Order.objects.create(
        order_number="0317-0001",
        client=client,
        device_type="Принтер",
        issue_description="X",
        received_date="2026-03-17",
        status=accepted,
        created_by=admin,
    )

    api = APIClient()
    _auth(api, email=admin.email, password="Passw0rd123")
    resp = api.get("/api/reports/orders/?from=2026-03-01&to=2026-03-31")
    assert resp.status_code == 200
    assert len(resp.data["results"]) == 1

    xlsx = api.get("/api/reports/orders.xlsx?from=2026-03-01&to=2026-03-31")
    assert xlsx.status_code == 200
    wb = load_workbook(BytesIO(xlsx.content))
    assert "Заказы" in wb.sheetnames


@pytest.mark.django_db
def test_tc32_stock_movements_report_and_export():
    admin = User.objects.create_user(email="admin@example.com", name="Admin", role=User.Role.ADMIN, password="Passw0rd123")
    cat = InventoryCategory.objects.create(name="Картриджи")
    p = Product.objects.create(name="HP", category=cat, sku="S1", unit="шт", min_stock=Decimal("1"))
    stock_in(product=p, quantity=Decimal("5"), created_by=admin, comment="in")
    stock_out(product=p, quantity=Decimal("1"), created_by=admin, reason="write_off", comment="out")

    api = APIClient()
    _auth(api, email=admin.email, password="Passw0rd123")
    resp = api.get("/api/reports/stock-movements/")
    assert resp.status_code == 200
    assert len(resp.data["results"]) >= 2

    xlsx = api.get("/api/reports/stock-movements.xlsx")
    assert xlsx.status_code == 200
    wb = load_workbook(BytesIO(xlsx.content))
    assert "Движение" in wb.sheetnames


@pytest.mark.django_db
def test_finance_report_basic():
    admin = User.objects.create_user(email="admin@example.com", name="Admin", role=User.Role.ADMIN, password="Passw0rd123")
    master = User.objects.create_user(email="m@example.com", name="M", role=User.Role.MASTER, password="Passw0rd123")
    client = Client.objects.create(type="person", name="Иванов", phone="+70000000000")
    accepted = OrderStatus.objects.get(code="accepted")
    from orders.models import Order

    cat = InventoryCategory.objects.create(name="Картриджи")
    p = Product.objects.create(
        name="HP",
        category=cat,
        sku="S1",
        unit="шт",
        min_stock=Decimal("0"),
        purchase_price=Decimal("100"),
        selling_price=Decimal("100"),
    )
    stock_in(product=p, quantity=Decimal("5"), created_by=admin, comment="in")

    o = Order.objects.create(
        order_number="0317-0001",
        client=client,
        device_type="Принтер",
        issue_description="X",
        received_date="2026-03-17",
        status=accepted,
        created_by=admin,
        assigned_master=master,
        final_work_cost=Decimal("300"),
        final_materials_cost=Decimal("200"),
        total_amount=Decimal("500"),
        materials_cost_price=Decimal("200"),
        profit=Decimal("300"),
        final_cost=Decimal("500"),
    )
    from inventory.models import OrderUsedProduct

    OrderUsedProduct.objects.create(
        order=o,
        product=p,
        quantity=Decimal("2"),
        selling_price_at_moment=Decimal("100"),
        purchase_price_at_moment=Decimal("100"),
    )

    api = APIClient()
    _auth(api, email=admin.email, password="Passw0rd123")
    resp = api.get("/api/reports/finance/?from=2026-03-01&to=2026-03-31")
    assert resp.status_code == 200
    assert resp.data["totals"]["revenue"] == "500"
    assert resp.data["totals"]["cost"] == "200"


@pytest.mark.django_db
def test_finance_report_xlsx_export_contains_totals_row():
    admin = User.objects.create_user(email="admin@example.com", name="Admin", role=User.Role.ADMIN, password="Passw0rd123")
    master = User.objects.create_user(email="m@example.com", name="M", role=User.Role.MASTER, password="Passw0rd123")
    client = Client.objects.create(type="person", name="Иванов", phone="+70000000000")
    accepted = OrderStatus.objects.get(code="accepted")

    cat = InventoryCategory.objects.create(name="Картриджи")
    p = Product.objects.create(
        name="HP",
        category=cat,
        sku="S1",
        unit="шт",
        min_stock=Decimal("0"),
        purchase_price=Decimal("100"),
        selling_price=Decimal("100"),
    )
    stock_in(product=p, quantity=Decimal("5"), created_by=admin, comment="in")

    o = Order.objects.create(
        order_number="0317-0001",
        client=client,
        device_type="Принтер",
        issue_description="X",
        received_date="2026-03-17",
        status=accepted,
        created_by=admin,
        assigned_master=master,
        final_work_cost=Decimal("300"),
        final_materials_cost=Decimal("200"),
        total_amount=Decimal("500"),
        materials_cost_price=Decimal("200"),
        profit=Decimal("300"),
        final_cost=Decimal("500"),
    )
    from inventory.models import OrderUsedProduct

    OrderUsedProduct.objects.create(
        order=o,
        product=p,
        quantity=Decimal("2"),
        selling_price_at_moment=Decimal("100"),
        purchase_price_at_moment=Decimal("100"),
    )

    api = APIClient()
    _auth(api, email=admin.email, password="Passw0rd123")
    xlsx = api.get("/api/reports/finance.xlsx?from=2026-03-01&to=2026-03-31")
    assert xlsx.status_code == 200

    wb = load_workbook(BytesIO(xlsx.content))
    assert "Финансы" in wb.sheetnames
    ws = wb["Финансы"]

    # Last row should contain totals marker in first column.
    last_row = ws.max_row
    assert ws.cell(row=last_row, column=1).value == "ИТОГО"

