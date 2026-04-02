"""
Импорт товаров склада из Excel (.xlsx / .xls).

Поддерживаются:
- простая таблица: «Артикул», «Товар», «Количество», «Ед. изм.», «Цена»;
- печатная накладная (УПД/ТОРГ): шапка сверху, таблица с «Товары (работы, услуги)», «Кол-во»,
  разреженные колонки; подвал («Итого», «Всего наименований»…) не импортируется.

Цены вида «7 721,28» (пробелы + запятая) нормализуются в Decimal.
"""

from __future__ import annotations

import io
import logging
import re
from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import BinaryIO

from django.core.files.uploadedfile import UploadedFile
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# xlrd 1.2.x — чтение бинарного .xls (Excel 97–2003)
try:
    import xlrd  # type: ignore[import-untyped]
except ImportError:  # pragma: no cover
    xlrd = None


@dataclass
class ParsedRow:
    """Одна строка данных после парсинга листа (1-based номер строки в файле)."""

    sheet_row: int
    sku: str
    name: str
    quantity: Decimal
    unit: str
    purchase_price: Decimal | None


@dataclass
class ImportApplyResult:
    created_count: int = 0
    stock_in_count: int = 0
    skipped_rows: int = 0
    product_ids_touched: list[int] = field(default_factory=list)
    created_product_ids: list[int] = field(default_factory=list)
    row_errors: list[dict] = field(default_factory=list)
    parse_errors: list[str] = field(default_factory=list)


# --- нормализация заголовков и сопоставление колонок ---

# «номер» не включаем — в накладных часто колонка «№» п/п, её нельзя путать с артикулом.
_HEADER_ALIASES_SKU = frozenset(
    {"артикул", "код", "код товара", "sku", "артикул товара", "код номенклатуры"}
)
_HEADER_ALIASES_NAME = frozenset(
    {
        "товар",
        "товары",
        "наименование",
        "название",
        "номенклатура",
        "описание",
        "наименование товара",
        "товары (работы, услуги)",
    }
)
_HEADER_ALIASES_QTY = frozenset(
    {"количество", "кол-во", "кол", "колво", "остаток", "кол-во товара", "количество товара"}
)
_HEADER_ALIASES_UNIT = frozenset(
    {"единица измерения", "ед изм", "ед.изм.", "ед.", "ед", "единица", "ед измерения"}
)
_HEADER_ALIASES_PRICE = frozenset(
    {"цена", "закуп", "закупка", "закупочная", "закупочная цена", "цена закуп", "цена закупки"}
)

# Подвал накладной — дальше строки не разбираем как товары (даже если есть числа).
_FOOTER_MARKERS = (
    "итого",
    "всего наименований",
    "в том числе ндс",
    "всего к оплате",
    "всего отпущено",
    "всего с наименованием",
    "сумма ндс",
    " руб.",
)


def _normalize_header_cell(val: object) -> str:
    if val is None:
        return ""
    s = str(val).strip().lower().replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    return s


def _header_matches_name(normalized_header: str) -> bool:
    h = normalized_header
    if not h:
        return False
    if h in _HEADER_ALIASES_NAME:
        return True
    if h.startswith("товар"):
        return True
    if h.startswith("наименование"):
        return True
    return False


def _row_smells_like_footer(row: list[object], scan_cols: int = 55) -> bool:
    parts: list[str] = []
    limit = min(scan_cols, len(row))
    for i in range(limit):
        v = row[i]
        if v is None:
            continue
        s = str(v).strip().lower().replace("ё", "е")
        if s:
            parts.append(s)
    blob = " ".join(parts)
    for frag in _FOOTER_MARKERS:
        if frag in blob:
            return True
    return False


def _parse_decimal_cell(val: object) -> Decimal | None:
    if val is None:
        return None
    if isinstance(val, Decimal):
        return val
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        try:
            return Decimal(str(val))
        except InvalidOperation:
            return None
    s = str(val).strip()
    if not s:
        return None
    s = s.replace("\xa0", " ").replace(" ", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s in (".", "-", "-."):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _parse_quantity_cell(val: object) -> Decimal | None:
    q = _parse_decimal_cell(val)
    if q is None:
        return None
    if q <= 0:
        return None
    return q.quantize(Decimal("0.001"))


def _normalize_unit(raw: str) -> str:
    u = raw.strip().lower().replace("ё", "е")
    if not u:
        return "шт"
    aliases = {
        "шт.": "шт",
        "штук": "шт",
        "шт": "шт",
        "м.": "м",
        "м": "м",
        "метр": "м",
        "метров": "м",
        "кг.": "кг",
        "кг": "кг",
        "килограмм": "кг",
        "л.": "л",
        "л": "л",
        "литр": "л",
        "упак.": "упак",
        "упак": "упак",
        "упаковка": "упак",
        "компл.": "компл",
        "компл": "компл",
        "комплект": "компл",
    }
    return aliases.get(u, u[:16])


def _score_header_row(cells: list[str]) -> tuple[int, dict[str, int]]:
    """Возвращает (score, mapping field -> col_index)."""
    norm = [_normalize_header_cell(c) for c in cells]
    mapping: dict[str, int] = {}
    for i, h in enumerate(norm):
        if not h:
            continue
        if h in _HEADER_ALIASES_SKU and "sku" not in mapping:
            mapping["sku"] = i
        elif _header_matches_name(h) and "name" not in mapping:
            mapping["name"] = i
        elif h in _HEADER_ALIASES_QTY and "quantity" not in mapping:
            mapping["quantity"] = i
        elif h in _HEADER_ALIASES_UNIT and "unit" not in mapping:
            mapping["unit"] = i
        elif h in _HEADER_ALIASES_PRICE and "price" not in mapping:
            mapping["price"] = i
    score = len(mapping)
    return score, mapping


def _read_rows_openpyxl(data: BinaryIO) -> list[list[object]]:
    wb = load_workbook(data, read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _read_rows_xlrd(data: BinaryIO) -> list[list[object]]:
    if xlrd is None:
        raise RuntimeError("Для файлов .xls установите пакет xlrd==1.2.0")
    raw = data.read()
    book = xlrd.open_workbook(file_contents=raw)
    sheet = book.sheet_by_index(0)
    out: list[list[object]] = []
    for r in range(sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            v = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE and v:
                try:
                    t = xlrd.xldate_as_tuple(v, book.datemode)
                    v = f"{t[0]:04d}-{t[1]:02d}-{t[2]:02d}"
                except Exception:
                    pass
            row.append(v)
        out.append(row)
    return out


def detect_header_and_mapping(grid: list[list[object]], scan_limit: int = 60) -> tuple[int, dict[str, int]] | None:
    best: tuple[int, int, dict[str, int]] | None = None  # score, row_idx, mapping
    limit = min(scan_limit, len(grid))
    for ri in range(limit):
        row = grid[ri]
        if not row:
            continue
        cells = [row[i] if i < len(row) else "" for i in range(len(row))]
        score, mapping = _score_header_row([str(c) if c is not None else "" for c in cells])
        if (
            score >= 3
            and mapping.get("sku") is not None
            and mapping.get("name") is not None
            and mapping.get("quantity") is not None
        ):
            if best is None or score > best[0]:
                best = (score, ri, mapping)
    if best is None:
        return None
    return best[1], best[2]


def parse_uploaded_inventory_excel(upload: UploadedFile | BinaryIO, filename: str) -> tuple[list[ParsedRow], list[str]]:
    """
    Читает файл, находит строку заголовков, возвращает список ParsedRow и ошибки парсинга.
    """
    errors: list[str] = []
    name_l = filename.lower()
    if hasattr(upload, "read"):
        upload.seek(0)
        raw = upload.read()
        upload.seek(0)
    else:
        raw = upload.read()
    bio = io.BytesIO(raw)

    try:
        if name_l.endswith(".xlsx") or name_l.endswith(".xlsm"):
            grid = _read_rows_openpyxl(bio)
        elif name_l.endswith(".xls"):
            grid = _read_rows_xlrd(io.BytesIO(raw))
        else:
            return [], ["Поддерживаются файлы .xlsx и .xls"]
    except Exception as e:
        logger.exception("excel read failed")
        return [], [f"Не удалось прочитать файл: {e}"]

    detected = detect_header_and_mapping(grid)
    if not detected:
        return [], [
            "Не найдена строка заголовков с колонками «Артикул», «Товар» и «Количество» (дополнительно: ед. изм., цена)."
        ]

    header_row_idx, colmap = detected
    data_start = header_row_idx + 1
    out: list[ParsedRow] = []

    def cell(row: list[object], key: str) -> object:
        idx = colmap.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for ri in range(data_start, len(grid)):
        row = grid[ri]
        if not row or all(v is None or (isinstance(v, str) and not str(v).strip()) for v in row):
            continue
        if _row_smells_like_footer(row):
            break
        sheet_row = ri + 1  # 1-based для сообщений пользователю
        sku_raw = cell(row, "sku")
        name_raw = cell(row, "name")
        sku = str(sku_raw).strip() if sku_raw is not None else ""
        name = str(name_raw).strip() if name_raw is not None else ""
        if not sku and not name:
            continue
        if not sku:
            errors.append(f"Строка {sheet_row}: пустой артикул.")
            continue
        if not name:
            errors.append(f"Строка {sheet_row}: пустое наименование (артикул «{sku}»).")
            continue
        qty = _parse_quantity_cell(cell(row, "quantity"))
        if qty is None:
            errors.append(f"Строка {sheet_row}: неверное или отсутствует количество («{sku}»).")
            continue
        unit_raw = cell(row, "unit")
        unit = _normalize_unit(str(unit_raw).strip() if unit_raw is not None else "шт")
        price = _parse_decimal_cell(cell(row, "price"))
        if price is not None:
            price = price.quantize(Decimal("0.01"))
        out.append(
            ParsedRow(
                sheet_row=sheet_row,
                sku=sku,
                name=name,
                quantity=qty,
                unit=unit,
                purchase_price=price,
            )
        )

    return out, errors


def aggregate_parsed_rows(rows: list[ParsedRow]) -> list[ParsedRow]:
    """Суммирует количество для одинаковых (sku, price) внутри файла."""
    key_fn = lambda r: (r.sku.strip(), r.purchase_price)
    buckets: dict[tuple[str, Decimal | None], list[ParsedRow]] = defaultdict(list)
    for r in rows:
        buckets[key_fn(r)].append(r)
    merged: list[ParsedRow] = []
    for (_sku, _price), group in buckets.items():
        first = group[0]
        total_q = sum((x.quantity for x in group), Decimal("0"))
        merged.append(
            ParsedRow(
                sheet_row=first.sheet_row,
                sku=first.sku.strip(),
                name=first.name,
                quantity=total_q,
                unit=first.unit,
                purchase_price=first.purchase_price,
            )
        )
    return merged


def _prices_match_for_stock_in(excel_price: Decimal | None, db_price: Decimal | None) -> bool:
    """Поступление на существующую карточку: пустая цена в файле или в базе не считается конфликтом партий."""
    if excel_price is None:
        return True
    if db_price is None:
        return True
    return excel_price == db_price.quantize(Decimal("0.01"))


def _variant_sku(base_sku: str, price: Decimal) -> str:
    """Уникальный артикул для той же номенклатуры с другой закупочной ценой (поле sku max 64)."""
    suffix = f"-Z{price.quantize(Decimal('0.01'))}".replace(".", "_")
    max_base = 64 - len(suffix)
    if max_base < 1:
        suffix = suffix[-63:]
        max_base = 1
    base = base_sku.strip()[:max_base] if len(base_sku) > max_base else base_sku.strip()
    return base + suffix


def resolve_unique_variant_sku(base_sku: str, price: Decimal, exists) -> str:
    """exists(sku: str) -> bool — артикул ≤64 символов без коллизии в БД."""

    base_candidate = _variant_sku(base_sku, price)
    if len(base_candidate) > 64:
        base_candidate = base_candidate[:64]
    if not exists(base_candidate):
        return base_candidate
    for n in range(2, 10_000):
        suf = f"-{n}"
        stem = _variant_sku(base_sku, price)
        if len(stem) + len(suf) > 64:
            stem = stem[: 64 - len(suf)]
        cand = stem + suf
        if not exists(cand):
            return cand
    return base_candidate[:64]


def apply_parsed_rows_to_db(
    rows: list[ParsedRow],
    *,
    default_category,
    user,
) -> ImportApplyResult:
    """
    Бизнес-логика:
    - новый артикул → создание товара + поступление на склад (движение);
    - артикул есть, цена в файле пустая или совпадает с закупкой → только поступление (+количество);
    - артикул есть, в файле указана другая закупка → новая карточка с суффиксом артикула, без изменения старой.
    """
    from django.db import transaction

    from inventory.models import Product
    from inventory.services import stock_in

    result = ImportApplyResult()
    rows = aggregate_parsed_rows(rows)

    def touch(pid: int) -> None:
        if pid not in result.product_ids_touched:
            result.product_ids_touched.append(pid)

    with transaction.atomic():
        for pr in rows:
            sku_key = pr.sku.strip()
            try:
                existing = Product.objects.select_for_update().filter(sku=sku_key).first()
                if existing is None:
                    p = Product.objects.create(
                        name=pr.name[:255],
                        category=default_category,
                        sku=sku_key[:64],
                        unit=pr.unit[:16],
                        purchase_price=pr.purchase_price,
                        min_stock=Decimal("0"),
                        current_stock=Decimal("0"),
                    )
                    result.created_count += 1
                    stock_in(
                        product=p,
                        quantity=pr.quantity,
                        created_by=user,
                        comment="Импорт Excel",
                    )
                    result.stock_in_count += 1
                    touch(p.pk)
                    result.created_product_ids.append(p.pk)
                    continue

                if _prices_match_for_stock_in(pr.purchase_price, existing.purchase_price):
                    stock_in(
                        product=existing,
                        quantity=pr.quantity,
                        created_by=user,
                        comment="Импорт Excel",
                    )
                    result.stock_in_count += 1
                    if existing.purchase_price is None and pr.purchase_price is not None:
                        existing.purchase_price = pr.purchase_price
                        existing.save(update_fields=["purchase_price", "updated_at"])
                    touch(existing.pk)
                    continue

                if pr.purchase_price is None:
                    result.row_errors.append(
                        {
                            "row": pr.sheet_row,
                            "sku": sku_key,
                            "message": "В базе другая закупочная цена; укажите цену в файле для отдельной карточки.",
                        }
                    )
                    result.skipped_rows += 1
                    continue

                def sku_taken(s: str) -> bool:
                    return Product.objects.filter(sku=s).exists()

                new_sku = resolve_unique_variant_sku(sku_key, pr.purchase_price, sku_taken)
                p = Product.objects.create(
                    name=pr.name[:255],
                    category=default_category,
                    sku=new_sku,
                    unit=pr.unit[:16],
                    purchase_price=pr.purchase_price,
                    min_stock=Decimal("0"),
                    current_stock=Decimal("0"),
                )
                result.created_count += 1
                stock_in(
                    product=p,
                    quantity=pr.quantity,
                    created_by=user,
                    comment="Импорт Excel (другая закупка)",
                )
                result.stock_in_count += 1
                touch(p.pk)
                result.created_product_ids.append(p.pk)
            except Exception as e:
                logger.exception("import row failed")
                result.row_errors.append(
                    {"row": pr.sheet_row, "sku": sku_key, "message": str(e)}
                )
                result.skipped_rows += 1

    return result
